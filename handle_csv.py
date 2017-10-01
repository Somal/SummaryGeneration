import csv
import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, colors

PROBLEM_HASHTAG = '#косяк'
GOAL_DAY_HASHTAG = '#цельнадень'
GOAL_WEEK_HASHTAG = '#цельнанеделю'
N = 200
red = PatternFill(fgColor=colors.RED, fill_type='solid')
green = PatternFill(fgColor=colors.GREEN, fill_type='solid')


def set_value(ws, i, j, value):
    ws.cell(row=i + 1, column=j + 1).value = value


today = datetime.date.today()
first_date = today - datetime.timedelta(days=today.weekday())
week_days = [first_date]
for i in range(1, 7):
    week_days.append(first_date + datetime.timedelta(days=i))
today_str = str(today)
filename = "backups/TickTick-backup-{}.csv".format(today)
date_format = '%Y-%m-%dT%X%z'

# Delete unused info
f = open(filename, 'r+')
lines = f.readlines()
if lines[0].find('''"Date:''') > -1:
    lines = lines[6:]
    print('find')

f.close()

new_f = open(filename, 'w')
new_f.writelines(lines)
new_f.close()

# Open it's as csv
f = open(filename, 'r+')
csvreader = csv.reader(f, delimiter=',', quotechar='"')
fieldnames = next(csvreader)

completed = []
uncompleted = []
problems = []
goals_on_day = []
goals_on_week = []
csvdictreader = csv.DictReader(f, fieldnames=fieldnames)
for l in csvdictreader:
    status = l['Status'].strip()
    completed_time = l['Completed Time'].strip()
    order = int(l['Order'].strip())
    start_date = l['Start Date'].strip()
    title = l['Title'].strip()

    # Completed
    if completed_time.startswith(today_str) and status == '2':
        completed.append(l)
    elif title.find(PROBLEM_HASHTAG) >= 0 and status == '0':
        problems.append(l)
    elif title.find(GOAL_DAY_HASHTAG) >= 0:
        goals_on_day.append(l)
    elif title.find(GOAL_WEEK_HASHTAG) >= 0:
        goals_on_week.append(l)
    # UnCompleted
    elif status == '0' and len(start_date) > 0:
        start_date = datetime.datetime.strptime(start_date, date_format).date()
        if start_date <= today:
            uncompleted.append(l)

uncompleted = sorted(uncompleted, key=lambda x: x['Priority'], reverse=True)
completed = sorted(completed, key=lambda x: x['Priority'], reverse=True)
# output_dict = {'completed': completed, 'uncompleted': uncompleted, 'problems': problems, 'goals': goals}

# Excel
sum_filename = "summaries/Summary_of_week_{}-{}.xlsx".format(week_days[0], week_days[-1])
wb_existed = False
try:
    wb = load_workbook(sum_filename)
    wb_existed = True
except Exception as e:
    wb = Workbook()
    sheet_names = wb.get_sheet_names()
    for ws_names in sheet_names:
        wb.remove_sheet(wb.get_sheet_by_name(ws_names))
    for day in reversed(week_days):
        ws = wb.create_sheet(str(day), 0)
    wb.save(sum_filename)

ws = wb.get_sheet_by_name(today_str)
for i in range(N):
    for j in range(N):
        ws.cell(row=i + 1, column=j + 1).alignment = Alignment(horizontal='center', vertical='center')
set_value(ws, 0, 0, 'Цель на неделю')
set_value(ws, 0, 1, 'Цель на день')
set_value(ws, 0, 2, 'Косяки')
# ws.column_dimensions['A'].width = 20
for i, goal in enumerate(goals_on_week):
    set_value(ws, i + 1, 0, goal['Title'])

for i, goal in enumerate(goals_on_day):
    set_value(ws, i + 1, 0, goal['Title'])

for i, pr in enumerate(problems):
    set_value(ws, i + 1, 2, pr['Title'])
    ws.cell(row=i + 2, column=3).fill = green if pr['Status'] == '2' else red

task_col = 4
set_value(ws, 0, task_col, 'Выполненные задачи')
set_value(ws, 0, task_col + 1, 'Приоритеты')
set_value(ws, 0, task_col + 2, 'Невыполненные задачи')
set_value(ws, 0, task_col + 3, 'Приоритеты')

for i, c in enumerate(completed):
    set_value(ws, i + 1, task_col, c['Title'])
    set_value(ws, i + 1, task_col + 1, c['Priority'])

for i, uc in enumerate(uncompleted):
    set_value(ws, i + 1, task_col + 2, uc['Title'])
    set_value(ws, i + 1, task_col + 3, uc['Priority'])

retro_row = max(len(problems), max(len(goals_on_day), len(goals_on_week))) + 2
set_value(ws, retro_row, 0, '+')
set_value(ws, retro_row, 1, '-')
ws.merge_cells(start_row=retro_row + 1, start_column=3, end_row=retro_row + 1, end_column=4)
set_value(ws, retro_row, 2, 'Ретро')
ws.merge_cells(start_row=retro_row + 2, start_column=3, end_row=retro_row + 5, end_column=4)

# Add autocorrection of width
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
for col, value in dims.items():
    ws.column_dimensions[col].width = value + 1
wb.save(sum_filename)
wb.close()
